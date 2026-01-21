#!/usr/bin/env python3
"""
Extract specific test data from the new datasets
"""

import csv
import json

def extract_csv_test_data():
    """Extract specific test cases from CSV"""
    print("🔍 EXTRACTING TEST DATA FROM CSV")
    print("=" * 50)
    
    test_cases = []
    
    try:
        with open('datasets/customer_support_tickets.csv', 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            # Get first 10 tickets for testing
            for i, row in enumerate(reader):
                if i >= 10:
                    break
                
                test_case = {
                    'ticket_id': row['Ticket ID'],
                    'customer': row['Customer Name'],
                    'product': row['Product Purchased'],
                    'ticket_type': row['Ticket Type'],
                    'subject': row['Ticket Subject'],
                    'status': row['Ticket Status'],
                    'resolution': row['Resolution']
                }
                test_cases.append(test_case)
                
                print(f"Ticket {test_case['ticket_id']}: {test_case['customer']} - {test_case['product']}")
                print(f"  Type: {test_case['ticket_type']} | Status: {test_case['status']}")
                print(f"  Subject: {test_case['subject']}")
                if test_case['resolution']:
                    print(f"  Resolution: {test_case['resolution'][:100]}...")
                print()
        
        return test_cases
        
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        return []

def extract_excel_info():
    """Try to extract Excel info without pandas"""
    print("🔍 EXCEL FILE INFO")
    print("=" * 50)
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook('datasets/realistic_customer_orders_india.xlsx')
        ws = wb.active
        
        print(f"Excel file loaded successfully!")
        print(f"Worksheet: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {ws.max_column}")
        
        # Get headers
        headers = []
        for col in range(1, min(ws.max_column + 1, 20)):  # First 20 columns
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
        
        print(f"Headers: {headers}")
        
        # Get first few rows
        print("\nSample data:")
        for row in range(2, min(5, ws.max_row + 1)):  # First 3 data rows
            row_data = []
            for col in range(1, min(len(headers) + 1, 10)):  # First 10 columns
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            print(f"Row {row-1}: {row_data}")
        
        return headers
        
    except ImportError:
        print("❌ openpyxl not installed - cannot read Excel file")
        return []
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return []

if __name__ == "__main__":
    print("📊 EXTRACTING TEST DATA FROM NEW DATASETS")
    print("=" * 60)
    
    # Extract CSV data
    csv_data = extract_csv_test_data()
    
    # Extract Excel info
    excel_headers = extract_excel_info()
    
    # Save test data
    test_data = {
        'csv_tickets': csv_data,
        'excel_headers': excel_headers
    }
    
    with open('new_dataset_test_data.json', 'w') as f:
        json.dump(test_data, f, indent=2)
    
    print(f"\n✅ Test data extracted and saved to new_dataset_test_data.json")